from os import path, listdir, remove

from openpyxl import Workbook
from openpyxl import load_workbook

def add_row(ws, row_index, data):
    col_index = 1
    for info in data:
        try:
            ws.cell(row = row_index, column = col_index, value = float(info))
        except ValueError:
            ws.cell(row = row_index, column = col_index, value = info)
        col_index += 1

def add_student(_id, institution, student_class, headers, data):
    filename = path.join('data', _id, 'download', institution + '.xlsx')

    if path.isfile(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        wb['Sheet'].title = student_class
        add_row(wb[student_class], 1, headers)

    try:
        ws = wb[student_class]
    except KeyError:
        ws = wb.create_sheet(student_class)
        ws.title = student_class
        add_row(ws, 1, headers)

    add_row(ws, ws.max_row + 1, data)

    wb.save(filename)
