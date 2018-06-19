from sys import argv, stdin, stdout
from os import path
import json

from openpyxl import Workbook

def main(saveLocation):
    headers = stdin.readline()[:-1].split(',')

    wb = Workbook()
    wb.remove(wb.active)

    while True:
        line = stdin.readline()[:-1]
        if line == '':
            break

        student = json.loads(line)
        if not student['grade'] in wb.sheetnames:
            wb.create_sheet(student['grade'])
            wb[student['grade']].append(headers)
        
        wb[student['grade']].append(list(student.values()))
    
    wb.save(path.join(saveLocation, 'report.xlsx'))

    stdout.write('DONE excel_report')
    stdout.flush()

#start process
if __name__ == '__main__':
    if len(argv) >= 2:
        main(argv[1])
    else:
        stdout.write('ERROR')
        stdout.flush()